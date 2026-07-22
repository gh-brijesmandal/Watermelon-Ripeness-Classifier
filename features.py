"""
Watermelon ripeness acoustic pipeline:
  1. Load audio 
  2. 4th-order Butterworth bandpass filter (70-2000 Hz, zero-phase)
  3. Detect the tap onset (with missing-tap catching)
  4. Extract a 250 ms resonance window starting 3 ms after the tap
  5. Plot the tap + resonance window, and the resonance spectrum
  6. Extract ripeness-relevant features from the resonance window

  Usage: 
  python or python3 features.py --mass 5 --id 12 --file melon1_tap1.wav
  i.e, python features.py --mass (the mass value in grams) --id (melon id) --file (file name in audio folder)
"""

import os 
import numpy as np
import soundfile as sf
from scipy.signal import butter, filtfilt, hilbert
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import argparse

# take data from command line
parser = argparse.ArgumentParser()
parser.add_argument("--file", type=str, required=True)
parser.add_argument("--mass", type=float, required=False)  # make sure its in grams
parser.add_argument("--id", type=int, required=True)
args = parser.parse_args()

# Define your variables here
file_path = "./audio/" + args.file
mass = (args.mass) or (0)
melon_id = args.id
excel_path = "./Data/Features.xlsx"

# this is audio loading and tap detecting part
def load_audio(file_path):
    "Loads an audio file and returns (signal, sample_rate), or (None, None) if the file is corrupt/unreadable"
    try:
        signal, sr = sf.read(file_path)
        if signal.ndim > 1:
            signal = signal[:, 0]
        return signal, sr
    except Exception as e:
        print(f"Failed to load {file_path}: {e}")
        return None, None

def bandpass_filter(signal, sr, lowcut=70, highcut=2000, order=4):
    "Applies a 4th-order zero-phase Butterworth bandpass filter (70-2000 Hz) to remove noise"
    nyq = sr / 2
    b, a = butter(order, [lowcut / nyq, highcut / nyq], btype='band')
    return filtfilt(b, a, signal)

def detect_tap_onset(signal, sr, threshold_ratio=0.8):
    "Detects the tap onset as the first sample crossing an amplitude threshold, returns None if no tap is found"
    threshold = threshold_ratio * np.max(np.abs(signal))
    if threshold <= 0:
        print("No tap detected: signal has no significant amplitude")
        return None
    above = np.where(np.abs(signal) > threshold)[0]
    if len(above) == 0:
        print("No tap detected: threshold never crossed")
        return None
    return above[0]

def extract_resonance_window(signal, sr, onset_idx, delay_ms=3, window_ms=250):
    "Extracts a fixed-length resonance window starting a fixed delay after the detected tap onset"
    start = onset_idx + int(sr * delay_ms / 1000)       # delay ms is how many seconds after the tap do you want the resonance window 
    end = start + int(sr * window_ms / 1000)
    if end > len(signal):
        print("Resonance window exceeds signal length, truncating")
        end = len(signal)
    return signal[start:end]

def plot_tap_and_spectrum(signal, sr, onset_idx, window):
    "Plots the full tap waveform with the resonance window highlighted, plus the window's frequency spectrum"
    t_signal = np.arange(len(signal)) / sr
    win_start = onset_idx / sr
    t_window = win_start + np.arange(len(window)) / sr

    freqs = np.fft.rfftfreq(len(window), d=1 / sr)
    mags = np.abs(np.fft.rfft(window))

    fig, axes = plt.subplots(2, 1, figsize=(10, 6))
    axes[0].plot(t_signal, signal, color='gray', label='Full signal')
    axes[0].plot(t_window, window, color='red', label='Resonance window')
    axes[0].set_xlabel('Time (s)')
    axes[0].set_ylabel('Amplitude')
    axes[0].legend()

    axes[1].plot(freqs, mags)
    axes[1].set_xlabel('Frequency (Hz)')
    axes[1].set_ylabel('Magnitude')

    plt.tight_layout()
    plt.show()


# this is feature extraction part
def fundamental_resonance_frequency(signal, sr):
    "Finds the fundamental frequency via autocorrelation (periodicity-based, not FFT-peak-based)"
    corr = np.correlate(signal, signal, mode='full')
    corr = corr[len(corr) // 2:]
    d = np.diff(corr)
    start = np.where(d > 0)[0]
    if len(start) == 0:
        return 0.0
    start = start[0]
    peak = np.argmax(corr[start:]) + start
    if peak == 0:
        return 0.0
    return sr / peak

def peak_frequency(signal, sr):
    "Finds the frequency bin with the highest magnitude in the FFT spectrum"
    n = len(signal)
    freqs = np.fft.rfftfreq(n, d=1 / sr)
    mags = np.abs(np.fft.rfft(signal))
    return freqs[np.argmax(mags)]

def peak_amplitude(signal):
    "Returns the maximum absolute amplitude in the time-domain signal"
    return np.max(np.abs(signal))

def damping_coefficient(signal, sr):
    "Fits an exponential decay to the signal envelope to estimate damping rate"
    envelope = np.abs(hilbert(signal))
    envelope[envelope == 0] = 1e-12
    t = np.arange(len(signal)) / sr
    log_env = np.log(envelope)
    slope, _ = np.polyfit(t, log_env, 1)
    return -slope
 
def zero_crossing_rate(signal, sr):
    "Counts sign changes per second in the time-domain signal"
    crossings = np.where(np.diff(np.sign(signal)))[0]
    duration = len(signal) / sr
    return len(crossings) / duration
 

def spectral_centroid(signal, sr):
    "Computes the magnitude-weighted mean frequency of the spectrum"
    n = len(signal)
    freqs = np.fft.rfftfreq(n, d=1 / sr)
    mags = np.abs(np.fft.rfft(signal))
    if np.sum(mags) == 0:
        return 0.0
    return np.sum(freqs * mags) / np.sum(mags)
 
def spectral_rolloff(signal, sr, roll_percent=0.85):
    "Finds the frequency below which a given percentage of spectral energy is contained"
    n = len(signal)
    freqs = np.fft.rfftfreq(n, d=1 / sr)
    mags = np.abs(np.fft.rfft(signal))
    energy = mags ** 2
    cumulative = np.cumsum(energy)
    threshold = roll_percent * cumulative[-1]
    idx = np.searchsorted(cumulative, threshold)
    return freqs[min(idx, len(freqs) - 1)]

def stiffness_index(peak_freq, mass): 
    "This gives estimate of firmness quality of the watermelon and is highly correlated with ripeness."
    if (mass == 0): 
        return None
    return (peak_freq ** 2) * (mass ** (2/3))


def write_features_to_excel(file_path, data: dict):
    "Appends a row of feature data to the Excel file, creating it with headers if it doesn't exist"
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(list(data.keys()))
 
    ws.append(list(data.values()))
    wb.save(file_path)


def main():
    "This function is responsible for handling the flow of the program."
    audio, sr = load_audio(file_path)
    if (sr == None):
        raise SystemExit("Audio Loading Error. Terminating the Program.")

    audio = bandpass_filter(audio,sr)

    onset_idx = detect_tap_onset(audio, sr)
    if onset_idx is not None:
        window = extract_resonance_window(audio, sr, onset_idx)
        plot_tap_and_spectrum(audio, sr, onset_idx, window)
    else:
        print("Skipping file: no tap detected")
        raise SystemExit("Tap CouldNot be Detected. Terminating the Program.")

    # Take the features from here now. // btw window is the extracted resonance audio
    fund_res_freq = fundamental_resonance_frequency(window,sr)
    peak_freq = peak_frequency(window, sr)
    peak_amp = peak_amplitude(window)
    damp_coeff = damping_coefficient(window, sr)
    zcr = zero_crossing_rate(window,sr)
    spec_cent = spectral_centroid(window, sr)
    spec_roll = spectral_rolloff(window, sr)
    stiff_index = stiffness_index(peak_freq = peak_freq, mass = mass)

    feature_data = {
        "melon_id": melon_id,
        "mass": mass,
        "fund_res_freq": fund_res_freq,
        "peak_freq": peak_freq,
        "peak_amp": peak_amp,
        "damp_coeff": damp_coeff,
        "zcr": zcr,
        "spec_cent": spec_cent,
        "spec_roll": spec_roll,
        "stiff_index": stiff_index
    }

    print(f"""
            Mass: {mass}
            Fundamental Resonance Frequency: {fund_res_freq},
            Peak Frequency: {peak_freq},
            Peak Amplitude: {peak_amp},
            Damping Coefficient: {damp_coeff},
            Zero Crossing Rate: {zcr},
            Spectral Centroid: {spec_cent},
            Spectral Rolloff: {spec_roll},
            Stiffness Index: {stiff_index}.
    """)
    write_features_to_excel(excel_path, data=feature_data)


if __name__=="__main__":
    main()
